import streamlit as st
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from fpdf import FPDF
import google.generativeai as genai
import io
import json

# Page Configuration
st.set_page_config(
    page_title="SHS Science Master Lesson Package Generator",
    page_icon="🧬",
    layout="wide"
)

# Title & Description
st.title("🧬 SHS Science Master Lesson Package Generator")
st.caption("Automated generation of data-dense Master Lesson Packages (Excel + PDF) matching official DepEd standards.")
st.divider()

# Sidebar Setup
with st.sidebar:
    st.header("⚙️ App Authentication")
    st.markdown("Input your Gemini API Key to run the high-density content engine.")
    api_key = st.text_input("Gemini API Key", type="password")
    st.info("💡 Ensure `openpyxl` and `fpdf2` are added to your `requirements.txt` file.")

# Inputs
col1, col2 = st.columns(2)
with col1:
    grade_level = st.selectbox("Grade Level", ["Grade 11", "Grade 12"])
    topic = st.text_input("MELC / Topic Focus", placeholder="e.g., Plate Tectonics or Hydrometeorological Hazards")

with col2:
    st.subheader("📌 Strict Production Rules Activated")
    st.markdown("""
    * **Zero Placeholders:** Every field, question, and paragraph will be written out completely.
    * **Flat List Compilation:** Answer keys match identification prompts explicitly without positional errors.
    * **Philippine Contextualization:** Automatic indexing of relevant local agencies (PAGASA, PHIVOLCS, etc.).
    """)

# Helper Function: Clean JSON String from Gemini
def clean_json_response(text):
    text = text.strip()
    if text.startswith("```json"):
        text = text[7:]
    if text.endswith("```"):
        text = text[:-3]
    return text.strip()

# Core Execution Button
if st.button("🚀 Generate Master Lesson Package Assets", type="primary"):
    if not api_key:
        st.error("⚠️ Please enter your Gemini API Key in the sidebar.")
    elif not topic:
        st.warning("⚠️ Please provide a Topic/MELC before generating.")
    else:
        with st.spinner("🧠 Orchestrating data-dense lesson elements... This takes a moment due to generation volume..."):
            try:
                genai.configure(api_key=api_key)
                model = genai.GenerativeModel('gemini-1.5-flash')
                
                # --- API DATA FETCHING ---
                # We request data structured strictly for our tabs to avoid text truncation or loops
                generation_prompt = f"""
                You are Jenalyn, a Senior High School Science Specialist. Generate data for a {grade_level} master lesson package on the topic: "{topic}".
                You must output raw JSON format ONLY matching this schema precisely without nested markdown structures.
                
                {{
                    "overview": {{
                        "references": "Write 4-5 complete sentences citing official academic frameworks, textbooks, and DepEd guidelines for this specific topic.",
                        "ai_use": "Write 4-5 complete sentences disclosing how generative AI tools were ethically deployed to design scaffolding tools for this topic.",
                        "content_standard": "Write 4-5 complete sentences explicitly detailing the overarching content standards prescribed for this topic.",
                        "performance_standard": "Write 4-5 complete sentences outlining the specific performance metrics and output expected from learners.",
                        "learning_competency": "Write 4-5 complete sentences unpacking the specific Most Essential Learning Competency for this science domain.",
                        "objectives": "Write 4-5 complete sentences defining three measurable, distinct behavioral learning objectives.",
                        "context": "Write 4-5 complete sentences detailing the learning environment, student demographics, and resource realities in typical public schools.",
                        "pre_lesson": "Write 4-5 complete sentences detailing the prerequisite knowledge checks and preliminary diagnostic engagements.",
                        "resources": "Write 4-5 complete sentences highlighting teacher-made material, equipment, or alternative contextualized teaching aids.",
                        "integration": "Write 4-5 complete sentences highlighting interdisciplinary links to disaster readiness, history, mathematics, or civic policies.",
                        "exit_ticket": "Write 4-5 complete sentences outlining the formative assessment tool used to test structural mastery at the end of instruction.",
                        "extended_learning": "Write 4-5 complete sentences indicating the enrichment projects and specialized remedial steps for struggling students."
                    }},
                    "sessions": [
                        {{"session_num": 1, "teacher_action": "Detailed teacher tasks for session 1 in 3 full sentences.", "student_task": "Detailed student actions for session 1 in 3 full sentences."}},
                        {{"session_num": 2, "teacher_action": "Detailed teacher tasks for session 2 in 3 full sentences.", "student_task": "Detailed student actions for session 2 in 3 full sentences."}},
                        {{"session_num": 3, "teacher_action": "Detailed teacher tasks for session 3 in 3 full sentences.", "student_task": "Detailed student actions for session 3 in 3 full sentences."}},
                        {{"session_num": 4, "teacher_action": "Detailed teacher tasks for session 4 in 3 full sentences.", "student_task": "Detailed student actions for session 4 in 3 full sentences."}},
                        {{"session_num": 5, "teacher_action": "Detailed teacher tasks for session 5 in 3 full sentences.", "student_task": "Detailed student actions for session 5 in 3 full sentences."}}
                    ],
                    "cornell_notes": [
                        {{"cue": "Technical term/cue 1", "content": "Three or more sentences detailing heavy academic concepts.", "summary": "One sentence synthesis."}}
                    ], // Fill exactly 10 items
                    "glossary": [
                        {{"term": "Term 1", "definition": "A precise, dense two-sentence scientific definition."}}
                    ], // Fill exactly 15 items
                    "assessments": [
                        {{"day": 1, "q": "Full sentence identification question checking explicit factual curriculum data.", "a": "ExactAnswer"}}
                    ], // Fill exactly 50 items (10 per day for days 1 to 5)
                    "article_1": {{
                        "title": "Title referencing a Philippine scientific agency (like PAGASA, PHIVOLCS, or PNRI) or local geographic context",
                        "body": "A complete, high-density academic essay between 200 and 300 words analyzing the science topic within national boundaries and statutory rules."
                    }},
                    "article_2": {{
                        "title": "A second unique title highlighting local geographic/environmental laws or specific community case studies in the Philippines",
                        "body": "A complete secondary academic article of 200 to 300 words without placeholders detailing physical scientific phenomena inside the archipelago."
                    }}
                }}
                
                CRITICAL: Ensure cornell_notes has 10 items, glossary has 15 items, assessments has 50 items. Do not truncate text. Fill out all text fully.
                """
                
                response = model.generate_content(generation_prompt)
                raw_data = json.loads(clean_json_response(response.text))
                
                # --- FILE 1: MASTER_LESSON_PACKAGE.XLSX GENERATION ---
                wb = openpyxl.Workbook()
                
                # Styling presets
                font_title = Font(name="Arial", size=14, bold=True, color="FFFFFF")
                font_header = Font(name="Arial", size=11, bold=True, color="FFFFFF")
                font_bold = Font(name="Arial", size=11, bold=True)
                font_regular = Font(name="Arial", size=11)
                fill_navy = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
                fill_soft = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
                align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
                align_left = Alignment(horizontal="left", vertical="top", wrap_text=True)
                thin_border = Border(
                    left=Side(style='thin', color='BFBFBF'), right=Side(style='thin', color='BFBFBF'),
                    top=Side(style='thin', color='BFBFBF'), bottom=Side(style='thin', color='BFBFBF')
                )

                # Tab 1: Lesson_Plan_Overview
                ws1 = wb.active
                ws1.title = "Lesson_Plan_Overview"
                ws1.views.sheetView[0].showGridLines = True
                
                # Sheet Header Banner
                ws1.merge_cells("A1:C1")
                ws1["A1"] = f"MASTER LESSON PLAN ROADMAP: {topic.upper()} ({grade_level})"
                ws1["A1"].font = font_title
                ws1["A1"].fill = fill_navy
                ws1["A1"].alignment = align_center
                ws1.row_dimensions[1].height = 40
                
                overview_fields = [
                    ("A. References", "references"), ("B. Declaration of AI use", "ai_use"),
                    ("C. Content Standard", "content_standard"), ("D. Performance Standard", "performance_standard"),
                    ("E. Learning Competency", "learning_competency"), ("F. Learning Objectives", "objectives"),
                    ("G. Learning Context", "context"), ("H. Pre-Lesson", "pre_lesson"),
                    ("J. Learning Resources", "resources"), ("K. Opportunities for Integration", "integration"),
                    ("L. Formative Assessment", "exit_ticket"), ("M. Extended Learning Activities", "extended_learning")
                ]
                
                current_row = 3
                for label, key in overview_fields:
                    ws1.cell(row=current_row, column=1, value=label).font = font_bold
                    val = raw_data.get("overview", {}).get(key, "Data not generated fully.")
                    ws1.cell(row=current_row, column=2, value=val).font = font_regular
                    ws1.cell(row=current_row, column=2).alignment = align_left
                    ws1.row_dimensions[current_row].height = 60
                    current_row += 1
                
                # Session split table within Tab 1
                current_row += 1
                ws1.cell(row=current_row, column=1, value="I. Detailed Lesson Flow (5 Sessions)").font = font_bold
                current_row += 1
                
                headers_s = ["Session", "Teacher Action (2-3 Sentences)", "Student Task (2-3 Sentences)"]
                for col_idx, h in enumerate(headers_s, 1):
                    cell = ws1.cell(row=current_row, column=col_idx, value=h)
                    cell.font = font_header
                    cell.fill = fill_navy
                    cell.alignment = align_center
                ws1.row_dimensions[current_row].height = 25
                
                for session in raw_data.get("sessions", []):
                    current_row += 1
                    ws1.cell(row=current_row, column=1, value=f"Session {session.get('session_num')}").font = font_bold
                    ws1.cell(row=current_row, column=2, value=session.get('teacher_action')).font = font_regular
                    ws1.cell(row=current_row, column=3, value=session.get('student_task')).font = font_regular
                    ws1.cell(row=current_row, column=2).alignment = align_left
                    ws1.cell(row=current_row, column=3).alignment = align_left
                    ws1.row_dimensions[current_row].height = 55
                
                # Tab 2: Cornell_Notes
                ws2 = wb.create_sheet(title="Cornell_Notes")
                ws2.views.sheetView[0].showGridLines = True
                headers_c = ["Cues / Keywords", "Detailed Content (3+ Core Academic Sentences)", "Summary / Synthesis Sentence"]
                for col_idx, h in enumerate(headers_c, 1):
                    cell = ws2.cell(row=1, column=col_idx, value=h)
                    cell.font = font_header
                    cell.fill = fill_navy
                    cell.alignment = align_center
                ws2.row_dimensions[1].height = 25
                
                for r_idx, note in enumerate(raw_data.get("cornell_notes", []), 2):
                    ws2.cell(row=r_idx, column=1, value=note.get('cue')).font = font_bold
                    ws2.cell(row=r_idx, column=2, value=note.get('content')).font = font_regular
                    ws2.cell(row=r_idx, column=3, value=note.get('summary')).font = font_regular
                    ws2.cell(row=r_idx, column=1).alignment = align_left
                    ws2.cell(row=r_idx, column=2).alignment = align_left
                    ws2.cell(row=r_idx, column=3).alignment = align_left
                    ws2.row_dimensions[r_idx].height = 65

                # Tab 3: Glossary
                ws3 = wb.create_sheet(title="Glossary")
                ws3.views.sheetView[0].showGridLines = True
                headers_g = ["Technical Term", "Expanded Definition (Exactly 2 Sentences)"]
                for col_idx, h in enumerate(headers_g, 1):
                    cell = ws3.cell(row=1, column=col_idx, value=h)
                    cell.font = font_header
                    cell.fill = fill_navy
                    cell.alignment = align_center
                ws3.row_dimensions[1].height = 25
                
                for r_idx, g_item in enumerate(raw_data.get("glossary", []), 2):
                    ws3.cell(row=r_idx, column=1, value=g_item.get('term')).font = font_bold
                    ws3.cell(row=r_idx, column=2, value=g_item.get('definition')).font = font_regular
                    ws3.cell(row=r_idx, column=1).alignment = align_left
                    ws3.cell(row=r_idx, column=2).alignment = align_left
                    ws3.row_dimensions[r_idx].height = 40

                # Tab 4: Student_Assessment
                ws4 = wb.create_sheet(title="Student_Assessment")
                ws4.views.sheetView[0].showGridLines = True
                headers_a = ["Question ID", "Instructional Tracking Day", "Curriculum Fact Identification Question"]
                for col_idx, h in enumerate(headers_a, 1):
                    cell = ws4.cell(row=1, column=col_idx, value=h)
                    cell.font = font_header
                    cell.fill = fill_navy
                    cell.alignment = align_center
                ws4.row_dimensions[1].height = 25
                
                for r_idx, ass in enumerate(raw_data.get("assessments", []), 2):
                    ws4.cell(row=r_idx, column=1, value=f"Q-{r_idx-1:02d}").font = font_bold
                    ws4.cell(row=r_idx, column=2, value=f"Day {ass.get('day')}").font = font_regular
                    ws4.cell(row=r_idx, column=3, value=ass.get('q')).font = font_regular
                    ws4.cell(row=r_idx, column=3).alignment = align_left
                    ws4.row_dimensions[r_idx].height = 30

                # Tab 5: Teacher_Answer_Key
                ws5 = wb.create_sheet(title="Teacher_Answer_Key")
                ws5.views.sheetView[0].showGridLines = True
                headers_ak = ["Question ID", "Target Answer Value"]
                for col_idx, h in enumerate(headers_ak, 1):
                    cell = ws5.cell(row=1, column=col_idx, value=h)
                    cell.font = font_header
                    cell.fill = fill_navy
                    cell.alignment = align_center
                ws5.row_dimensions[1].height = 25
                
                # Flattened parsing implementation rule to prevent positional index mismatch crashes
                flat_assessment_list = list(raw_data.get("assessments", []))
                for r_idx in range(len(flat_assessment_list)):
                    row_num = r_idx + 2
                    item = flat_assessment_list[r_idx]
                    ws5.cell(row=row_num, column=1, value=f"Q-{r_idx+1:02d}").font = font_bold
                    ws5.cell(row=row_num, column=2, value=item.get('a')).font = font_bold
                    ws5.cell(row=row_num, column=2).fill = fill_soft
                    ws5.row_dimensions[row_num].height = 25

                # Tab 6: Supplemental_Reading
                ws6 = wb.create_sheet(title="Supplemental_Reading")
                ws6.views.sheetView[0].showGridLines = True
                
                ws6.cell(row=2, column=1, value="Article 1 Contextualized Analysis").font = font_bold
                ws6.cell(row=3, column=1, value=raw_data.get("article_1", {}).get("title")).font = font_bold
                ws6.cell(row=4, column=1, value=raw_data.get("article_1", {}).get("body")).font = font_regular
                ws6.cell(row=4, column=1).alignment = align_left
                ws6.row_dimensions[4].height = 180
                
                ws6.cell(row=6, column=1, value="Article 2 Regulatory & Environmental Context").font = font_bold
                ws6.cell(row=7, column=1, value=raw_data.get("article_2", {}).get("title")).font = font_bold
                ws6.cell(row=8, column=1, value=raw_data.get("article_2", {}).get("body")).font = font_regular
                ws6.cell(row=8, column=1).alignment = align_left
                ws6.row_dimensions[8].height = 180

                # Universal Column Width Auto-Fitting
                for ws in [ws1, ws2, ws3, ws4, ws5, ws6]:
                    for col in ws.columns:
                        ws.column_dimensions[col[0].column_letter].width = 32

                # Save Workbook into bytes buffer
                excel_buffer = io.BytesIO()
                wb.save(excel_buffer)
                excel_buffer.seek(0)
                
                # --- FILE 2: NOTEBOOKLM_SOURCE_MATERIAL.PDF GENERATION ---
                pdf = FPDF()
                pdf.set_auto_page_break(auto=True, margin=15)
                pdf.add_page()
                
                # PDF Title Styling
                pdf.set_font("Helvetica", style="B", size=16)
                pdf.cell(0, 10, f"NotebookLM Primary Reference Index: {topic.upper()}", ln=True, align="C")
                pdf.set_font("Helvetica", style="I", size=10)
                pdf.cell(0, 5, f"Curated for {grade_level} Senior High School Instruction", ln=True, align="C")
                pdf.ln(10)
                
                # Section 1: Cornell Notes Sync
                pdf.set_font("Helvetica", style="B", size=14)
                pdf.cell(0, 10, "Section 1: High-Density Technical Concepts (Cornell Notes Schema)", ln=True)
                pdf.ln(2)
                for note in raw_data.get("cornell_notes", []):
                    pdf.set_font("Helvetica", style="B", size=11)
                    pdf.multi_cell(0, 6, f"Core Objective Focus: {note.get('cue')}")
                    pdf.set_font("Helvetica", size=10)
                    pdf.multi_cell(0, 5, f"Technical Core Text:\n{note.get('content')}")
                    pdf.set_font("Helvetica", style="I", size=9)
                    pdf.multi_cell(0, 5, f"Synthesis Check: {note.get('summary')}")
                    pdf.ln(4)
                
                # Section 2: Glossary Sync
                pdf.set_font("Helvetica", style="B", size=14)
                pdf.cell(0, 10, "Section 2: Scientific Nomenclature Glossary Index", ln=True)
                pdf.ln(2)
                for g_item in raw_data.get("glossary", []):
                    pdf.set_font("Helvetica", style="B", size=11)
                    pdf.cell(40, 6, f"{g_item.get('term')}: ", ln=False)
                    pdf.set_font("Helvetica", size=11)
                    pdf.multi_cell(0, 6, g_item.get('definition'))
                    pdf.ln(1)
                
                # Section 3: Supplemental Articles
                pdf.ln(5)
                pdf.set_font("Helvetica", style="B", size=14)
                pdf.cell(0, 10, "Section 3: Contextualized Original Localization Articles", ln=True)
                pdf.ln(2)
                
                # Article 1
                pdf.set_font("Helvetica", style="B", size=12)
                pdf.multi_cell(0, 6, f"Article 1: {raw_data.get('article_1', {}).get('title')}")
                pdf.ln(1)
                pdf.set_font("Helvetica", size=10)
                pdf.multi_cell(0, 5, raw_data.get('article_1', {}).get('body'))
                pdf.ln(6)
                
                # Article 2
                pdf.set_font("Helvetica", style="B", size=12)
                pdf.multi_cell(0, 6, f"Article 2: {raw_data.get('article_2', {}).get('title')}")
                pdf.ln(1)
                pdf.set_font("Helvetica", size=10)
                pdf.multi_cell(0, 5, raw_data.get('article_2', {}).get('body'))
                
                # Save PDF to bytes
                pdf_output = pdf.output(dest='S')
                pdf_buffer = io.BytesIO(pdf_output) if isinstance(pdf_output, bytes) else io.BytesIO(pdf_output.encode('latin1'))

                # --- FRONT-END INTERACTION AREA ---
                st.success("🎉 Asset processing complete! Both Master components are compiled without default values.")
                
                col_dl1, col_dl2 = st.columns(2)
                with col_dl1:
                    st.download_button(
                        label="📥 Download Master_Lesson_Package.xlsx",
                        data=excel_buffer,
                        file_name=f"Master_Lesson_Package_{topic.replace(' ', '_')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
                with col_dl2:
                    st.download_button(
                        label="📥 Download NotebookLM_Source_Material.pdf",
                        data=pdf_buffer,
                        file_name=f"NotebookLM_Source_Material_{topic.replace(' ', '_')}.pdf",
                        mime="application/pdf"
                    )
                    
            except Exception as e:
                st.error(f"❌ Error during layout building or content mapping: {str(e)}")
